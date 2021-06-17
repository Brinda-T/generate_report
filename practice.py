from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, PatternFill, Protection, Alignment
from openpyxl.styles import NamedStyle
import mymodule
import datetime

sno_pos = mymodule.get_sno_pos()
print ("sno_pos", sno_pos)
issue_key_pos = mymodule.get_issue_key_pos()
print ("issue_key_pos", issue_key_pos)
issue_key_col_name = mymodule.get_issue_key_col_name()
print ("issue_key_col_name", issue_key_col_name)
issue_id_pos = mymodule.get_issue_id_pos()
print ("issue_id_pos", issue_id_pos)
issue_id_col_name = mymodule.get_issue_id_col_name()
print ("issue_id_col_name", issue_id_col_name)
custom_field_epiclink_pos = mymodule.get_custom_field_epiclink_pos()
print ("custom_field_epiclink_pos", custom_field_epiclink_pos)
custom_field_epiclink_col_name = mymodule.get_custom_field_epiclink_col_name()
print ("custom_field_epiclink_col_name", custom_field_epiclink_col_name)
ename_pos = mymodule.get_ename_pos()
print ("ename_pos", ename_pos)
ename_col_name = mymodule.get_ename_col_name()
print ("ename_col_name", ename_col_name)
assignee_pos = mymodule.get_assignee_pos()
print ("assignee_pos", assignee_pos)
assignee_col_name = mymodule.get_assignee_col_name()
print ("assignee_col_name", assignee_col_name)
custom_field_storypoints_pos = mymodule.get_custom_field_storypoints_pos()
print ("custom_field_storypoints_pos", custom_field_storypoints_pos)
custom_field_storypoints_col_name = mymodule.get_custom_field_storypoints_col_name()
print ("custom_field_storypoints_col_name", custom_field_storypoints_col_name)
teste_pos = mymodule.get_teste_pos()
print ("teste_pos", teste_pos)
teste_col_name = mymodule.get_teste_col_name()
print ("teste_col_name", teste_col_name)
original_estimate_pos = mymodule.get_original_estimate_pos()
print ("original_estimate_pos", original_estimate_pos)
original_estimate_col_name = mymodule.get_original_estimate_col_name()
print ("original_estimate_col_name", original_estimate_col_name)
time_spent_pos = mymodule.get_time_spent_pos()
print ("time_spent_pos", time_spent_pos)
time_spent_col_name = mymodule.get_time_spent_col_name()
print ("time_spent_col_name", time_spent_col_name)
remaining_estimate_pos = mymodule.get_remaining_estimate_pos()
print ("remaining_estimate_pos", remaining_estimate_pos)
remaining_estimate_col_name = mymodule.get_remaining_estimate_col_name()
print ("remaining_estimate_col_name", remaining_estimate_col_name)
sprint_pos = mymodule.get_sprint_pos()
print ("sprint_pos", sprint_pos)
sprint_col_name = mymodule.get_sprint_col_name()
print ("sprint_col_name", sprint_col_name)
sprint2_pos = mymodule.get_sprint2_pos()
print ("sprint2_pos", sprint2_pos)
sprint2_col_name = mymodule.get_sprint2_col_name()
print ("sprint2_col_name", sprint2_col_name)
sprint3_pos = mymodule.get_sprint3_pos()
print ("sprint3_pos", sprint3_pos)
sprint3_col_name = mymodule.get_sprint3_col_name()
print ("sprint3_col_name", sprint3_col_name)
summary_pos = mymodule.get_summary_pos()
print ("summary_pos", summary_pos)
summary_col_name = mymodule.get_summary_col_name()
print ("summary_col_name", summary_col_name)

print ("------------------")

epics_custom_field_pos = mymodule.get_epics_custom_field_pos()
print ("epics_custom_field_pos", epics_custom_field_pos)	
epics_Esdate_pos = mymodule.get_epics_Esdate_pos()
print ("epics_Esdate_pos", epics_Esdate_pos)
epics_Etdate_pos = mymodule.get_epics_Etdate_pos()
print ("epics_Etdate_pos", epics_Etdate_pos)
epics_EASdate_pos = mymodule.get_epics_EASdate_pos()
print ("epics_EASdate_pos", epics_EASdate_pos)
epics_EAEdate_pos = mymodule.get_epics_EAEdate_pos()
print ("epics_EAEdate_pos", epics_EAEdate_pos)

print ("--------------------------")

progress_pos = mymodule.get_progress_pos()
print ("progress_pos", progress_pos)
progress_gen_column_name = mymodule.get_progress_gen_column_name()
print ("progress_gen_column_name", progress_gen_column_name)

scheduled_progress_pos = mymodule.get_scheduled_progress_pos()
print ("scheduled_progress_pos", scheduled_progress_pos)
scheduled_progress_gen_column_name = mymodule.get_scheduled_progress_gen_column_name()
print ("scheduled_progress_gen_column_name", scheduled_progress_gen_column_name)

scheduled_overrun_pos = mymodule.get_scheduled_overrun_pos()
print ("scheduled_overrun_pos", scheduled_overrun_pos)
scheduled_overrun_gen_column_name = mymodule.get_scheduled_overrun_gen_column_name()
print ("scheduled_overrun_gen_column_name", scheduled_overrun_gen_column_name)

remarks_pos = mymodule.get_remarks_pos()
print ("remarks_pos", remarks_pos)
remarks_gen_column_name = mymodule.get_remarks_gen_column_name()
print ("remarks_gen_column_name", remarks_gen_column_name)


def print_all_worksheet_names(wbname):
	wb = load_workbook(wbname)

	print (wb.sheetnames)

	for wsheet in wb.sheetnames:
		print (wsheet)
	print (" ")

def print_row_column1(wbname,src_wname1,src_wname2,dst_wname):

	dates_list = []
	sprint_list = []

	wb = load_workbook(wbname)

	swsheet1 = wb.get_sheet_by_name(src_wname1)
	swsheet2 = wb.get_sheet_by_name(src_wname2)

	try:
		dwsheet = wb.get_sheet_by_name(dst_wname)
		print ("worksheet '%s' found"%(dst_wname))
		print ("removing worksheet:'%s'"%(dst_wname))
		wb.remove_sheet(dwsheet)
	except:
		print ("worksheet '%s' not found"%(dst_wname))
	finally:
		print ("creating new worksheet:'%s'"%(dst_wname))
		dwsheet = wb.create_sheet(dst_wname,0) 

	srcount1 = swsheet1.max_row
	sccount1 = swsheet1.max_column
	print ("%s: max row:col (%d:%d)"%(src_wname1,srcount1,sccount1))

	srcount2 = swsheet2.max_row
	sccount2 = swsheet2.max_column
	print ("%s: max row:col (%d:%d)"%(src_wname2,srcount2,sccount2))

	#drcount =  dwsheet.max_row
	#dccount =  dwsheet.max_column
	#print("%s: max row:col (%d:%d)" % (dst_wname, drcount, dccount))

	row = swsheet1[1]	
	frow = [cell.value for cell in row]
	dwsheet.append(frow)

	titles_center_aligned_text = Alignment(horizontal = "center", vertical = "center", wrapText=True)
	center_aligned_text = Alignment(horizontal = "center", vertical = "center")

	'''	
	dwsheet['A1'] = "sai\ndheeraj"
	dwsheet['A1'].alignment = Alignment(horizontal = "center", vertical = "center", wrapText=True)
	#dwsheet['A1'].alignment = center_aligned_text
	wb.save("wrap.xlsx")
	'''

	header_row = dwsheet[1]
	dwsheet[1][issue_key_pos].value = issue_key_col_name
	dwsheet[1][issue_key_pos].alignment = titles_center_aligned_text

	dwsheet[1][issue_id_pos].value = issue_id_col_name
	dwsheet[1][issue_id_pos].alignment = titles_center_aligned_text

	dwsheet[1][custom_field_epiclink_pos].value = custom_field_epiclink_col_name
	dwsheet[1][custom_field_epiclink_pos].alignment = titles_center_aligned_text

	dwsheet[1][ename_pos].value = ename_col_name
	dwsheet[1][ename_pos].alignment = titles_center_aligned_text

	dwsheet[1][assignee_pos].value = assignee_col_name
	dwsheet[1][assignee_pos].alignment = titles_center_aligned_text

	dwsheet[1][custom_field_storypoints_pos].value = custom_field_storypoints_col_name
	dwsheet[1][custom_field_storypoints_pos].alignment = titles_center_aligned_text

	dwsheet[1][teste_pos].value = teste_col_name
	dwsheet[1][teste_pos].alignment = titles_center_aligned_text

	dwsheet[1][original_estimate_pos].value = original_estimate_col_name
	dwsheet[1][original_estimate_pos].alignment = titles_center_aligned_text

	dwsheet[1][time_spent_pos].value = time_spent_col_name
	dwsheet[1][time_spent_pos].alignment = titles_center_aligned_text

	dwsheet[1][remaining_estimate_pos].value = remaining_estimate_col_name
	dwsheet[1][remaining_estimate_pos].alignment = titles_center_aligned_text

	dwsheet[1][sprint_pos].value = sprint_col_name
	dwsheet[1][sprint_pos].alignment = titles_center_aligned_text

	dwsheet[1][sprint2_pos].value = sprint2_col_name
	dwsheet[1][sprint2_pos].alignment = titles_center_aligned_text

	dwsheet[1][sprint3_pos].value = sprint3_col_name
	dwsheet[1][sprint3_pos].alignment = titles_center_aligned_text

	dwsheet[1][summary_pos].value = summary_col_name
	dwsheet[1][summary_pos].alignment = titles_center_aligned_text

	dwsheet['p1'] = progress_gen_column_name
	dwsheet['p1'].alignment = titles_center_aligned_text

	dwsheet['q1'] = scheduled_progress_gen_column_name
	dwsheet['q1'].alignment = titles_center_aligned_text

	dwsheet['r1'] = scheduled_overrun_gen_column_name
	dwsheet['r1'].alignment = titles_center_aligned_text

	dwsheet['s1'] = remarks_gen_column_name
	dwsheet['s1'].alignment = titles_center_aligned_text

	for j in range (2, srcount1+1):
		sprint_list.append(swsheet1[j][sprint_pos].value)
		sprint_list.append(swsheet1[j][sprint2_pos].value)
		sprint_list.append(swsheet1[j][sprint3_pos].value)
		sprint_list = list(filter(None, sprint_list))
		sprint_list.sort(reverse = True)
		#print (sprint_list)
		dwsheet[j][issue_id_pos].value = sprint_list[0]
		sprint_list = []
	
	for i in range (2, srcount1+1):
		dwsheet[i][sno_pos].value = swsheet1[i][sno_pos].value
		dwsheet[i][sno_pos].alignment = center_aligned_text

		dwsheet[i][issue_key_pos].value = swsheet1[i][custom_field_epiclink_pos].value
		dwsheet[i][issue_key_pos].alignment = center_aligned_text

		dwsheet[i][custom_field_epiclink_pos].value = swsheet1[i][issue_key_pos].value
		dwsheet[i][custom_field_epiclink_pos].alignment = center_aligned_text
		
		dwsheet[i][ename_pos].value = swsheet1[i][summary_pos].value
		dwsheet[i][ename_pos].alignment = center_aligned_text

		dwsheet[i][assignee_pos].value = swsheet1[i][assignee_pos].value
		dwsheet[i][assignee_pos].alignment = center_aligned_text

		dwsheet[i][custom_field_storypoints_pos].value = swsheet1[i][custom_field_storypoints_pos].value
		dwsheet[i][custom_field_storypoints_pos].alignment = center_aligned_text
		
		estimate_effort_hours = int(swsheet1[i][original_estimate_pos].value) / 3600
		dwsheet[i][time_spent_pos].value = estimate_effort_hours
		dwsheet[i][time_spent_pos].alignment = center_aligned_text

		dwsheet[i][remaining_estimate_pos].value = swsheet1[i][time_spent_pos].value
		dwsheet[i][remaining_estimate_pos].alignment = center_aligned_text

		effort_consumed_hours = int(swsheet1[i][remaining_estimate_pos].value) / 3600
		dwsheet[i][sprint_pos].value = effort_consumed_hours
		dwsheet[i][sprint_pos].alignment = center_aligned_text

		pending_effort_hours = estimate_effort_hours - effort_consumed_hours
		dwsheet[i][sprint2_pos].value = pending_effort_hours
		dwsheet[i][sprint2_pos].alignment = center_aligned_text

		effort_completion = (effort_consumed_hours/estimate_effort_hours) * 100
		dwsheet[i][progress_pos].value = ('{}{}'.format(round(effort_completion),"%"))
		dwsheet[i][progress_pos].alignment = center_aligned_text

		scheduled_progress = 100
		dwsheet[i][scheduled_progress_pos].value = ('{}{}'.format(scheduled_progress,"%"))
		dwsheet[i][scheduled_progress_pos].alignment = center_aligned_text

		scheduled_overrun = 0
		dwsheet[i][scheduled_overrun_pos].value = ('{}{}'.format(scheduled_overrun,"%"))
		dwsheet[i][scheduled_overrun_pos].alignment = center_aligned_text


	for k in range (2, srcount2+1):
		for l in range (2, srcount1+1):
			dates_list.append(swsheet2[k][epics_Esdate_pos].value)
			dates_list.append(swsheet2[k][epics_Etdate_pos].value)
			dates_list.append(swsheet2[k][epics_EASdate_pos].value)
			dates_list.append(swsheet2[k][epics_EAEdate_pos].value)
			dates_list.sort()

			if (swsheet2[k][epics_custom_field_pos].value == swsheet1[l][custom_field_epiclink_pos].value):

				dwsheet[l][teste_pos].value = datetime.datetime.date(dates_list[0])
				dwsheet[l][teste_pos].alignment = center_aligned_text
				#print (dates_list[0])
				#print (type(dates_list[0]))

				dwsheet[l][original_estimate_pos].value = datetime.datetime.date(dates_list[2])
				dwsheet[l][original_estimate_pos].alignment = center_aligned_text
				
				dwsheet[l][sprint3_pos].value = datetime.datetime.date(dates_list[1])
				dwsheet[l][sprint3_pos].alignment = center_aligned_text

				dwsheet[l][summary_pos].value = datetime.datetime.date(dates_list[3])
				dwsheet[l][summary_pos].alignment = center_aligned_text
				#print (dwsheet[l][original_estimate_pos].value)
			


	drcount =  dwsheet.max_row
	dccount =  dwsheet.max_column
	print("%s: max row:col (%d:%d)" % (dst_wname, drcount, dccount))

	print("Saving :%s" % (wbname))
	wb.save("wrap.xlsx")
	wb.save(wbname)


def main():
	filename = "employee-details.xlsx"
	src_wname1 = "02-stories"
	src_wname2 = "01-epics"
	dst_wname = "generated"
	print_all_worksheet_names(filename)
	print_row_column1(filename,src_wname1,src_wname2,dst_wname)

if (__name__ == '__main__'):
	main()