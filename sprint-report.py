from get_functions import create_workbook, get_read_csv_files, create_worksheet
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, PatternFill, Protection, Alignment
from openpyxl.styles import NamedStyle
from columns import get_column_names_and_positions
from sr_log import sr_log_messages, sr_debug

sno_pos, issue_key_pos, issue_key_col_name, issue_id_pos, issue_id_col_name, custom_field_epiclink_pos, custom_field_epiclink_col_name, ename_pos, ename_col_name, assignee_pos, assignee_col_name, custom_field_storypoints_pos, custom_field_storypoints_col_name,teste_pos, teste_col_name, original_estimate_pos, original_estimate_col_name, time_spent_pos, time_spent_col_name,remaining_estimate_pos, remaining_estimate_col_name, sprint_pos, sprint_col_name, sprint2_pos, sprint2_col_name,sprint3_pos, sprint3_col_name, summary_pos, summary_col_name, epics_custom_field_pos, epics_Esdate_pos, epics_Etdate_pos, epics_EASdate_pos, epics_EAEdate_pos, progress_pos, progress_gen_column_name, scheduled_progress_pos, scheduled_progress_gen_column_name, scheduled_overrun_pos, scheduled_overrun_gen_column_name, remarks_pos, remarks_gen_column_name = get_column_names_and_positions()

def get_col_names(wbname,dst_wname):  
    wb = load_workbook(wbname)
    dwsheet = wb[dst_wname]
    
    dwsheet.insert_cols(1, 18)
    header_row = dwsheet[1]
    dwsheet[1][issue_key_pos].value = issue_key_col_name
    dwsheet[1][issue_id_pos].value = issue_id_col_name
    dwsheet[1][custom_field_epiclink_pos].value = custom_field_epiclink_col_name
    dwsheet[1][ename_pos].value = ename_col_name
    dwsheet[1][assignee_pos].value = assignee_col_name
    dwsheet[1][custom_field_storypoints_pos].value = custom_field_storypoints_col_name
    dwsheet[1][teste_pos].value = teste_col_name
    dwsheet[1][original_estimate_pos].value = original_estimate_col_name
    dwsheet[1][time_spent_pos].value = time_spent_col_name
    dwsheet[1][remaining_estimate_pos].value = remaining_estimate_col_name
    dwsheet[1][sprint_pos].value = sprint_col_name
    dwsheet[1][sprint2_pos].value = sprint2_col_name
    dwsheet[1][sprint3_pos].value = sprint3_col_name
    dwsheet[1][summary_pos].value = summary_col_name
    dwsheet['p1'] = progress_gen_column_name
    dwsheet['q1'] = scheduled_progress_gen_column_name
    dwsheet['r1'] = scheduled_overrun_gen_column_name
    dwsheet['s1'] = remarks_gen_column_name

    wb.save(wbname)

def get_sprint_value(wbname, stories, dst_wname):
    
    sprint_list = []
    wb = load_workbook(wbname)
    dwsheet = wb[dst_wname]
	
    srcount1 = len(stories)   
    for j in range (1, srcount1):
        i = j + 1
        
        sprint_list.append(stories[j][sprint_pos])
        sprint_list.append(stories[j][sprint2_pos])
        sprint_list.append(stories[j][sprint3_pos])
        sprint_list = list(filter(None, sprint_list))
        sprint_list.sort(reverse = True)
        dwsheet[i][issue_id_pos].value = sprint_list[0]
        sprint_list = []
    wb.save(wbname)
def get_values_for_columns(wbname,stories,epics,dst_wname):
    dates_list = []
   	
    wb = load_workbook(wbname)
    dwsheet = wb[dst_wname]
	
    srcount1 = len(stories)   
    sr_debug ("stories length :%d"%(srcount1))

    srcount2 = len(epics)  
    sr_debug ("epics length :%d"%(srcount2))
    for i in range (1, srcount1):
        j = i + 1
        dwsheet[j][sno_pos].value = stories[i][sno_pos]
        dwsheet[j][issue_key_pos].value = stories[i][custom_field_epiclink_pos]
        dwsheet[j][custom_field_epiclink_pos].value = stories[i][issue_key_pos]
        dwsheet[j][ename_pos].value = stories[i][summary_pos]
        dwsheet[j][assignee_pos].value = stories[i][assignee_pos]
        dwsheet[j][custom_field_storypoints_pos].value = stories[i][custom_field_storypoints_pos]
        
        estimate_effort_hours = int(stories[i][original_estimate_pos]) / 3600
        dwsheet[j][time_spent_pos].value = estimate_effort_hours

        dwsheet[j][remaining_estimate_pos].value = stories[i][time_spent_pos]

        effort_consumed_hours = int(stories[i][time_spent_pos]) / 3600
        dwsheet[j][sprint_pos].value = effort_consumed_hours

        pending_effort_hours = estimate_effort_hours - effort_consumed_hours
        dwsheet[j][sprint2_pos].value = pending_effort_hours
        dwsheet['M3'] = "=J3-L3"

        effort_completion = (effort_consumed_hours/estimate_effort_hours) * 100
        dwsheet[j][progress_pos].value = ('{}{}'.format(round(effort_completion),"%"))
        dwsheet['P2'].value = "=L3/J3*100" 

        scheduled_progress = 100
        dwsheet[j][scheduled_progress_pos].value = ('{}{}'.format(scheduled_progress,"%"))

        scheduled_overrun = 0
        dwsheet[j][scheduled_overrun_pos].value = ('{}{}'.format(scheduled_overrun,"%"))
	
    wb.save(wbname)
def get_dates_append(wbname, stories, epics, dst_wname):
    dates_list = []
   	
    wb = load_workbook(wbname)
    dwsheet = wb[dst_wname]
	
    srcount1 = len(stories)   
    sr_debug ("stories length :%d"%(srcount1))

    srcount2 = len(epics)  
    sr_debug ("epics length :%d"%(srcount2))    
    for k in range (1, srcount2):
        for l in range (1, srcount1):
            i = l + 1
            dates_list.append(epics[k][epics_Esdate_pos])
            dates_list.append(epics[k][epics_Etdate_pos])
            dates_list.append(epics[k][epics_EASdate_pos])
            dates_list.append(epics[k][epics_EAEdate_pos])

            if (epics[k][epics_custom_field_pos]):
                stories[l][custom_field_epiclink_pos]

            dwsheet[i][teste_pos].value = (dates_list[0])
            dwsheet[i][original_estimate_pos].value = (dates_list[2])
            dwsheet[i][sprint3_pos].value = (dates_list[1])
            dwsheet[i][summary_pos].value = (dates_list[3])
 
    sr_debug("Saving :%s" % (wbname))
    wb.save("employee-details.xlsx")
    wb.save(wbname)

def get_cell_alignment(wbname, dst_wname):
    wb = load_workbook(wbname)
    dwsheet = wb[dst_wname]

    text_alignment_for_column = Alignment(horizontal = "center", vertical = "center", wrapText=True)
    text_alignment_for_row = Alignment(horizontal = "center", vertical = "center")
	
    rcount = dwsheet.max_row
    ccount = dwsheet.max_column
	
    for i in range(2, rcount+1):
        for j in range(0, ccount):
            dwsheet[1][j].alignment = text_alignment_for_column
            dwsheet[i][j].alignment = text_alignment_for_row
    wb.save(wbname)

def get_cell_colors_using_patternfill(wbname, wsheet):
    wb = load_workbook(wbname)
    dwsheet = wb[wsheet]

    rcount = dwsheet.max_row
    ccount = dwsheet.max_column
    
    fill_pattern = PatternFill(patternType = 'solid', fgColor = 'CCCCFF')
    for i in range(0, ccount):
        dwsheet[1][i].fill = fill_pattern
    
    wb.save(wbname)

def text_alignment(wbname, dst_wname):
    wb = load_workbook(wbname)
    dwsheet = wb[dst_wname]

    row_count = dwsheet.max_row
    column_count = dwsheet.max_column

    titles_center_aligned_text = Alignment(horizontal = "center", vertical = "center", wrapText=True)
    center_aligned_text = Alignment(horizontal = "center", vertical = "center")

    for i in range (2, row_count+1):
        for j in range (0, column_count):
            dwsheet[1][j].alignment = titles_center_aligned_text
            dwsheet[i][j].alignment = center_aligned_text

    wb.save(wbname)

def get_heading(wbname,wsheet):
    wb = load_workbook(wbname)
    dwsheet = wb[wsheet]
    
    min = dwsheet.min_column
    max = dwsheet.max_column
    dwsheet.insert_rows(1)
    dwsheet['A1'].value = "SPRINT/STORY BOARD"
    dwsheet.merge_cells('A1:S1')
    dwsheet['A1'].alignment = Alignment(horizontal = "center", vertical = "center") 
    
    fill_pattern = PatternFill(patternType = 'solid', fgColor = 'FFFF00')
    dwsheet['A1'].fill = fill_pattern
    wb.save(wbname)

def get_cell_formulae(wbname, dst_wname):
    wb = load_workbook(wbname)
    dwsheet = wb[dst_wname]
    dwsheet['M3'].value = "=J3-L3"
    dwsheet['P3'].value = "=J3/L3*100"
    drcount = dwsheet.max_row
    dccount = dwsheet.max_column
    for i in range(3, drcount + 1):
        dwsheet['L' + str(i)].value = "=" + "K" + str(i) + "/" + "3600"
        dwsheet['M' + str(i)].value = "=" + "J" + str(i) + "-" + "L" + str(i) 
        dwsheet['P' + str(i)].value = "=" + "L" + str(i) + "/" + "J" + str(i) + "%" 
        
    wb.save(wbname)


def main():
    filename = "employee-details.xlsx"
    stories = "02-stories.csv"
    epics = "01-epics.csv"
    dst_wname = "generated"
    create_workbook(filename)
    epics_data = get_read_csv_files(epics)
    stories_data = get_read_csv_files(stories)
    create_worksheet(filename, dst_wname)
    get_col_names(filename, dst_wname)
    get_sprint_value(filename, stories_data, dst_wname)
    get_values_for_columns(filename, stories_data, epics_data, dst_wname)
    get_dates_append(filename, stories_data, epics_data, dst_wname)
    get_cell_alignment(filename, dst_wname)
    get_cell_colors_using_patternfill(filename, dst_wname)
    get_heading(filename, dst_wname)
    get_cell_formulae(filename, dst_wname)
	
if (__name__ == '__main__'):
    main()

